"""AlpaTrade agent/tool eval harness (DeepEval, XAI judge).

Reads evals/ground_truth.csv, gets an answer from the right agent for each prompt,
grades it (deterministic → structural match; chat → DeepEval GEval via grok), and
writes CSV + XLSX reports with per-agent / per-type accuracy.

Usage:
    python evals/run_evals.py                    # chat + deterministic (skips slow)
    python evals/run_evals.py --include-slow     # also run agent:backtest / agent:paper
    python evals/run_evals.py --only chat        # one type only
    python evals/run_evals.py --limit 5          # first N rows
    python evals/run_evals.py --dry-run          # list prompts, no calls
"""
from __future__ import annotations

import argparse
import asyncio
import csv
import sys
import time
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
load_dotenv(ROOT / ".env")

GROUND_TRUTH = ROOT / "evals" / "ground_truth.csv"
RESULTS_DIR = ROOT / "eval-results"

_geval = None  # lazy GEval metric (grok judge)


def load_rows() -> list[dict]:
    with open(GROUND_TRUTH, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


# --------------------------------------------------------------------------- answers
async def get_answer(row: dict) -> str:
    """Route the prompt to the right agent and return its answer text."""
    prompt = row["prompt"]
    import agui_app  # builds the LangGraph agent + command interceptor

    if row["agent_type"] == "chat":
        from langchain_core.messages import HumanMessage
        r = await agui_app.langgraph_agent.ainvoke({"messages": [HumanMessage(content=prompt)]})
        msg = r["messages"][-1]
        return getattr(msg, "content", "") or ""

    # deterministic → run as a CLI command through the same interceptor the app uses
    from engine.ai import StreamingCommand
    result = await agui_app._command_interceptor(prompt, {"thread_id": "eval"})
    if result is None:
        return "(command not recognised)"
    if isinstance(result, StreamingCommand):
        from tui.command_processor import CommandProcessor
        cp = CommandProcessor(result.app_state, user_id=None)
        return await cp.process_command(result.raw_command) or ""
    return str(result)


# --------------------------------------------------------------------------- grading
def grade_deterministic(ai: str, must_contain: str):
    """Structural match. `must_contain` = groups joined by '|'; a group is ';'-separated
    alternatives. PASS when every group has at least one alternative present (case-insensitive)."""
    ai_l = (ai or "").lower()
    if ai_l.startswith("error:") or ai_l.startswith("(command not"):
        return False, 0.0, ai[:120]
    if not (must_contain or "").strip():
        ok = bool(ai_l.strip())
        return ok, (1.0 if ok else 0.0), "non-empty output" if ok else "empty output"
    groups = [g for g in must_contain.split("|") if g.strip()]
    missing = []
    for g in groups:
        alts = [a.strip().lower() for a in g.split(";") if a.strip()]
        if not any(a in ai_l for a in alts):
            missing.append(g.split(";")[0])
    ok = not missing
    score = round(1.0 - len(missing) / max(len(groups), 1), 3)
    return ok, score, ("all required tokens present" if ok else f"missing: {missing}")


def grade_chat(prompt: str, expected: str, ai: str):
    """GEval correctness graded by grok."""
    global _geval
    from deepeval.test_case import LLMTestCase
    from evals.judge import correctness_metric
    if _geval is None:
        _geval = correctness_metric(threshold=0.6)
    tc = LLMTestCase(input=prompt, actual_output=(ai or "(empty)"), expected_output=expected)
    _geval.measure(tc)
    ok = _geval.score is not None and _geval.score >= _geval.threshold
    return ok, round(_geval.score or 0.0, 3), (_geval.reason or "")[:300]


# --------------------------------------------------------------------------- run
async def run(rows: list[dict], include_slow: bool) -> list[dict]:
    results = []
    for row in rows:
        if row.get("slow") and not include_slow:
            continue
        t0 = time.time()
        try:
            ai = await get_answer(row)
        except Exception as e:  # noqa: BLE001
            ai = f"ERROR: {e}"
        latency = round(time.time() - t0, 2)
        if row["agent_type"] == "chat":
            ok, score, reason = grade_chat(row["prompt"], row["expected_answer"], ai)
        else:
            ok, score, reason = grade_deterministic(ai, row.get("must_contain", ""))
        results.append({
            "prompt": row["prompt"],
            "expected_answer": row["expected_answer"],
            "ai_answer": (ai or "").replace("\n", " ")[:900],
            "agent_name": row["agent_name"],
            "agent_type": row["agent_type"],
            "result": "PASS" if ok else "FAIL",
            "score": score,
            "reason": reason,
            "latency_s": latency,
        })
        print(f"  [{len(results):>2}] {'PASS' if ok else 'FAIL'} "
              f"({score:>4}) [{row['agent_type'][:4]}/{row['agent_name']}] {row['prompt'][:48]}")
    return results


# --------------------------------------------------------------------------- report
def _accuracy(results, key=None, val=None):
    subset = [r for r in results if key is None or r[key] == val]
    n = len(subset)
    p = sum(1 for r in subset if r["result"] == "PASS")
    return p, n, (100.0 * p / n if n else 0.0)


COLS = ["prompt", "expected_answer", "ai_answer", "agent_name", "agent_type",
        "result", "score", "reason", "latency_s"]


def save_and_report(results: list[dict]) -> dict:
    RESULTS_DIR.mkdir(exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S")

    csv_path = RESULTS_DIR / f"evals-{ts}.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=COLS)
        w.writeheader()
        w.writerows(results)

    # per-agent + per-type summaries
    by_agent = sorted({r["agent_name"] for r in results})
    agent_rows = [(a, *_accuracy(results, "agent_name", a)) for a in by_agent]
    type_rows = [(t, *_accuracy(results, "agent_type", t)) for t in ("chat", "deterministic")]
    p, n, acc = _accuracy(results)

    xlsx_path = RESULTS_DIR / f"evals-{ts}.xlsx"
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        wb = Workbook()
        ws = wb.active
        ws.title = "results"
        ws.append([c.replace("_", " ").title() for c in COLS])
        for c in ws[1]:
            c.font = Font(bold=True)
        green = PatternFill("solid", fgColor="C6EFCE")
        red = PatternFill("solid", fgColor="FFC7CE")
        for r in results:
            ws.append([r[c] for c in COLS])
            cell = ws.cell(row=ws.max_row, column=COLS.index("result") + 1)
            cell.fill = green if r["result"] == "PASS" else red
        for col, width in zip("ABCDEFGHI", (34, 40, 60, 20, 14, 8, 8, 40, 10)):
            ws.column_dimensions[col].width = width
        sm = wb.create_sheet("summary")
        sm.append(["Scope", "Passed", "Total", "Accuracy %"])
        sm["A1"].font = sm["B1"].font = sm["C1"].font = sm["D1"].font = Font(bold=True)
        sm.append(["OVERALL", p, n, round(acc, 1)])
        sm.append([])
        sm.append(["By type", "", "", ""])
        for t, tp, tn, ta in type_rows:
            sm.append([t, tp, tn, round(ta, 1)])
        sm.append([])
        sm.append(["By agent", "", "", ""])
        for a, ap, an, aa in agent_rows:
            sm.append([a, ap, an, round(aa, 1)])
        for col, width in zip("ABCD", (26, 10, 10, 12)):
            sm.column_dimensions[col].width = width
        wb.save(xlsx_path)
    except Exception as e:  # noqa: BLE001
        print(f"(xlsx skipped: {e})")
        xlsx_path = None

    print("\n" + "=" * 60)
    print(f"ACCURACY  {p}/{n}  =  {acc:.1f}%")
    print("-" * 60)
    for t, tp, tn, ta in type_rows:
        if tn:
            print(f"  {t:<14} {tp:>3}/{tn:<3}  {ta:5.1f}%")
    print("-" * 60)
    for a, ap, an, aa in agent_rows:
        print(f"  {a:<24} {ap:>3}/{an:<3}  {aa:5.1f}%")
    print("=" * 60)
    print(f"CSV : {csv_path}")
    if xlsx_path:
        print(f"XLSX: {xlsx_path}")
    return {"passed": p, "total": n, "accuracy": round(acc, 1)}


def main():
    ap = argparse.ArgumentParser(description="AlpaTrade agent/tool evals")
    ap.add_argument("--include-slow", action="store_true", help="also run slow agents (backtest/paper)")
    ap.add_argument("--only", choices=["chat", "deterministic"], help="run one agent_type only")
    ap.add_argument("--limit", type=int, default=None, help="first N rows")
    ap.add_argument("--dry-run", action="store_true", help="list prompts, no calls")
    args = ap.parse_args()

    rows = load_rows()
    if args.only:
        rows = [r for r in rows if r["agent_type"] == args.only]
    if args.limit:
        rows = rows[: args.limit]

    if args.dry_run:
        for i, r in enumerate(rows):
            slow = " [slow]" if r.get("slow") else ""
            print(f"  [{i:>2}] {r['agent_type'][:4]}/{r['agent_name']}: {r['prompt']}{slow}")
        print(f"\n{len(rows)} prompts.")
        return

    print(f"Running {len(rows)} evals (include_slow={args.include_slow})...\n")
    results = asyncio.run(run(rows, args.include_slow))
    save_and_report(results)


if __name__ == "__main__":
    main()
